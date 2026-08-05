"""
Módulo de transformación OCN (Orden de Compra a Proveedor).

Convierte un reporte de solicitudes de compra (misma estructura que el
usado para SCC) en el archivo plano OCN, cruzando automáticamente el
proveedor y el precio unitario desde la lista de precios, y la sucursal /
condición de pago desde el maestro de proveedores.
"""

import io

import pandas as pd
import openpyxl
from openpyxl.styles import Font

from transformations.utils import strip, format_fecha, to_number, normalizar_item, nit_base

REQUIRED_SOLICITUD_COLUMNS = [
    "C.O. docto.",
    "C.O. movto.",
    "Bodega",
    "U.M.",
    "Cant. solicitada",
    "Fecha solic.",
    "Item",
    "Estado",
]
REQUIRED_PROVEEDOR_COLUMNS = ["Nit", "Sucursal", "Condicion de pago", "Numero identificación"]
REQUIRED_PRECIO_COLUMNS = ["numero identificacion proveedor", "Item", "U.M.", "Precio unit."]


class OCNError(Exception):
    pass


def _validar_columnas(df, requeridas, nombre):
    faltantes = [c for c in requeridas if c not in df.columns]
    if faltantes:
        raise OCNError(f"El archivo de {nombre} no tiene las columnas esperadas. Faltan: {', '.join(faltantes)}")


def generar_ocn(
    df_solicitud: pd.DataFrame,
    df_proveedores: pd.DataFrame,
    df_precios: pd.DataFrame,
    tercero_comprador: str = "901906381",
) -> tuple[bytes, dict]:
    df = df_solicitud.copy()
    df.columns = [str(c).strip() for c in df.columns]
    _validar_columnas(df, REQUIRED_SOLICITUD_COLUMNS, "solicitudes de compra")

    dfp = df_proveedores.copy()
    dfp.columns = [str(c).strip() for c in dfp.columns]
    _validar_columnas(dfp, REQUIRED_PROVEEDOR_COLUMNS, "maestro de proveedores")

    dfl = df_precios.copy()
    dfl.columns = [str(c).strip() for c in dfl.columns]
    _validar_columnas(dfl, REQUIRED_PRECIO_COLUMNS, "lista de precios")

    total_filas = len(df)

    # --- Filtrar solo Aprobado ---
    if "Estado" in df.columns:
        df["Estado"] = df["Estado"].apply(strip)
        df = df[df["Estado"].str.lower() == "aprobado"].reset_index(drop=True)
    descartadas_estado = total_filas - len(df)

    if df.empty:
        raise OCNError("No quedó ninguna fila con Estado = 'Aprobado' en el archivo de solicitudes.")

    # Normalizar campos clave de la solicitud
    for col in ["C.O. docto.", "C.O. movto.", "Bodega", "U.M."]:
        df[col] = df[col].apply(strip)
    df["Item"] = df["Item"].apply(normalizar_item)
    df["Fecha solic."] = df["Fecha solic."].apply(format_fecha)
    if "Fecha entrega" in df.columns:
        df["Fecha entrega"] = df["Fecha entrega"].apply(format_fecha)
    df["Cant. solicitada"] = df["Cant. solicitada"].apply(to_number)

    # --- Maestro de precios: (Item, U.M.) -> (proveedor_base, precio) ---
    dfl["Item"] = dfl["Item"].apply(normalizar_item)
    dfl["U.M."] = dfl["U.M."].apply(strip)
    dfl["numero identificacion proveedor"] = dfl["numero identificacion proveedor"].apply(nit_base)
    precio_lookup = {}
    for _, r in dfl.iterrows():
        key = (r["Item"], r["U.M."])
        precio_lookup[key] = (r["numero identificacion proveedor"], to_number(r["Precio unit."]))

    # --- Maestro de proveedores: nit_base -> (nit_sin_guion, sucursal, condicion_pago) ---
    dfp["Numero identificación"] = dfp["Numero identificación"].apply(nit_base)
    dfp["Nit"] = dfp["Nit"].apply(strip)
    dfp["Sucursal"] = dfp["Sucursal"].apply(strip)
    dfp["Condicion de pago"] = dfp["Condicion de pago"].apply(strip)
    proveedor_lookup = {}
    for _, r in dfp.iterrows():
        proveedor_lookup[r["Numero identificación"]] = (r["Nit"], r["Sucursal"], r["Condicion de pago"])

    # --- Resolver proveedor + precio para cada fila ---
    filas_validas = []
    filas_sin_cruce = []
    for _, row in df.iterrows():
        item = row["Item"]
        um = row["U.M."]
        key = (item, um)
        if key not in precio_lookup:
            filas_sin_cruce.append(
                {"motivo": "Item/U.M. no encontrado en lista de precios", "Item": item, "U.M.": um, "Bodega": row["Bodega"]}
            )
            continue
        nit_base_val, precio = precio_lookup[key]
        if nit_base_val not in proveedor_lookup:
            filas_sin_cruce.append(
                {"motivo": "Proveedor no encontrado en maestro de proveedores", "Item": item, "U.M.": um, "Proveedor": nit_base_val}
            )
            continue
        nit_sin_guion, sucursal, condicion_pago = proveedor_lookup[nit_base_val]
        filas_validas.append(
            {
                "C.O. docto.": row["C.O. docto."],
                "C.O. movto.": row["C.O. movto."],
                "Bodega": row["Bodega"],
                "U.M.": um,
                "Cantidad": row["Cant. solicitada"],
                "Fecha solic.": row["Fecha solic."],
                "Fecha entrega": (row["Fecha entrega"] if "Fecha entrega" in df.columns and row["Fecha entrega"] else row["Fecha solic."]),
                "Item": item,
                "Precio": precio,
                "Proveedor": nit_sin_guion,
                "Sucursal": sucursal,
                "C Pago": condicion_pago,
            }
        )

    if not filas_validas:
        raise OCNError(
            "Ninguna fila del archivo de solicitudes pudo cruzarse con la lista de precios / maestro de proveedores."
        )

    resolved = pd.DataFrame(filas_validas)

    # --- Consecutivo: por (centro, proveedor), numerado independiente por centro ---
    consecutivo_por_centro = {}
    grupo_consecutivo = {}
    orden_grupos = []
    for _, row in resolved.iterrows():
        gkey = (row["C.O. docto."], row["Proveedor"])
        if gkey not in grupo_consecutivo:
            centro = row["C.O. docto."]
            consecutivo_por_centro[centro] = consecutivo_por_centro.get(centro, 0) + 1
            grupo_consecutivo[gkey] = consecutivo_por_centro[centro]
            orden_grupos.append(gkey)

    wb = openpyxl.Workbook()
    ws_docs = wb.active
    ws_docs.title = "Documentos"
    ws_docs.append(["C.O", "TIPO DCTO", "NRO CONSECUTIVO", "FECHA", "TERCERO COMPRADOR", "TERCERO PROVEEDOR", "SUCURSAL", "C PAGO"])

    documentos_generados = 0
    for gkey in orden_grupos:
        centro, proveedor = gkey
        sub = resolved[(resolved["C.O. docto."] == centro) & (resolved["Proveedor"] == proveedor)]
        consecutivo = grupo_consecutivo[gkey]
        fecha = sub["Fecha solic."].iloc[0]
        sucursal = sub["Sucursal"].iloc[0]
        c_pago = sub["C Pago"].iloc[0]
        ws_docs.append([centro, "OCN", consecutivo, fecha, tercero_comprador, proveedor, sucursal, c_pago])
        documentos_generados += 1

    ws_mov = wb.create_sheet("Movimientos")
    ws_mov.append(
        ["CO", "TIPO DCTO", "NRO CONSECUTIVO", "NRO REGISTRO", "BODEGA", "C.O MOVIMIENTO", "UM", "CANTIDAD", "AAAAMMDD ENTREGA", "PRECIO UNITARIO", "ITEM"]
    )
    movimientos_generados = 0
    for gkey in orden_grupos:
        centro, proveedor = gkey
        sub = resolved[(resolved["C.O. docto."] == centro) & (resolved["Proveedor"] == proveedor)]
        consecutivo = grupo_consecutivo[gkey]
        for i, (_, row) in enumerate(sub.iterrows(), start=1):
            ws_mov.append(
                [
                    centro,
                    "OCN",
                    consecutivo,
                    i,
                    row["Bodega"],
                    row["C.O. movto."],
                    row["U.M."],
                    row["Cantidad"],
                    row["Fecha entrega"],
                    row["Precio"],
                    row["Item"],
                ]
            )
            movimientos_generados += 1

    # Forzar texto en columna ITEM para no perder ceros a la izquierda
    item_col_idx = 11
    for row in ws_mov.iter_rows(min_row=2, min_col=item_col_idx, max_col=item_col_idx):
        for cell in row:
            cell.number_format = "@"

    for ws in (ws_docs, ws_mov):
        for cell in ws[1]:
            cell.font = Font(bold=True)

    # Hoja de errores/pendientes, si los hay
    if filas_sin_cruce:
        ws_err = wb.create_sheet("No cruzados")
        ws_err.append(["Motivo", "Item", "U.M.", "Bodega", "Proveedor"])
        for e in filas_sin_cruce:
            ws_err.append([e.get("motivo", ""), e.get("Item", ""), e.get("U.M.", ""), e.get("Bodega", ""), e.get("Proveedor", "")])
        for cell in ws_err[1]:
            cell.font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resumen = {
        "total_filas_entrada": total_filas,
        "descartadas_por_estado": descartadas_estado,
        "filas_sin_cruce": len(filas_sin_cruce),
        "documentos_generados": documentos_generados,
        "movimientos_generados": movimientos_generados,
    }
    return buf.read(), resumen
