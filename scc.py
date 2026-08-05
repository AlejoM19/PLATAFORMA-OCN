"""
Módulo de transformación SCC (Solicitud de Consumo/Traslado interno).

Convierte un reporte de solicitudes aprobadas (una fila por ítem solicitado)
en el archivo plano SCC con hojas Inicial / Documentos / Movimientos / Final.
"""

import io

import pandas as pd
import openpyxl
from openpyxl.styles import Font

from transformations.utils import strip, format_fecha, to_number, normalizar_item

REQUIRED_COLUMNS = [
    "C.O. docto.",
    "C.O. movto.",
    "Bodega",
    "Motivo",
    "C.Costo",
    "U.M.",
    "Cant. solicitada",
    "Fecha solic.",
    "Item",
    "Estado",
]


class SCCError(Exception):
    pass


def validar_columnas(df):
    faltantes = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if faltantes:
        raise SCCError(
            "El archivo de solicitudes no tiene las columnas esperadas. "
            f"Faltan: {', '.join(faltantes)}"
        )


def generar_scc(df_solicitud: pd.DataFrame, notas: str) -> tuple[bytes, dict]:
    """
    Genera el archivo plano SCC a partir del reporte de solicitudes.

    Devuelve (bytes_del_excel, resumen) donde resumen trae estadísticas
    útiles para mostrar en la interfaz (documentos generados, filas
    descartadas, etc).
    """
    df = df_solicitud.copy()
    df.columns = [str(c).strip() for c in df.columns]
    validar_columnas(df)

    total_filas = len(df)

    # Solo filas Aprobado
    df["Estado"] = df["Estado"].apply(strip)
    df = df[df["Estado"].str.lower() == "aprobado"].reset_index(drop=True)
    descartadas_estado = total_filas - len(df)

    if df.empty:
        raise SCCError(
            "No quedó ninguna fila con Estado = 'Aprobado' en el archivo de solicitudes."
        )

    # Normalizar campos clave
    df["C.O. docto."] = df["C.O. docto."].apply(strip)
    df["C.O. movto."] = df["C.O. movto."].apply(strip)
    df["Bodega"] = df["Bodega"].apply(strip)
    df["Motivo"] = df["Motivo"].apply(to_number)
    df["C.Costo"] = df["C.Costo"].apply(strip)
    df["U.M."] = df["U.M."].apply(strip)
    df["Item"] = df["Item"].apply(normalizar_item)
    df["Fecha solic."] = df["Fecha solic."].apply(format_fecha)
    df["Cant. solicitada"] = df["Cant. solicitada"].apply(to_number)

    # Cada CENTRO DE OPERACION (C.O. docto.) es un documento independiente.
    # (Una misma "Nro solic." puede agrupar varios centros distintos en el
    # reporte de origen, pero el archivo plano exige un documento por centro).
    grupos_orden = list(dict.fromkeys(df["C.O. docto."]))
    consecutivo_por_grupo = {co: i + 1 for i, co in enumerate(grupos_orden)}

    wb = openpyxl.Workbook()

    # --- Hoja Inicial ---
    ws_inicial = wb.active
    ws_inicial.title = "Inicial"
    ws_inicial.append(["COMPANIA"])
    ws_inicial.append([1])

    # --- Hoja Documentos ---
    ws_docs = wb.create_sheet("Documentos")
    ws_docs.append(
        ["COMPANIA", "CENTRO DE OPERACION", "CONSECUTIVO", "FECHA  AAAAMMDD", "NOTAS", "DCTO REFERENCIA"]
    )
    documentos_generados = 0
    for centro in grupos_orden:
        sub = df[df["C.O. docto."] == centro]
        fecha = sub["Fecha solic."].iloc[0]
        consecutivo = consecutivo_por_grupo[centro]
        ws_docs.append([1, centro, consecutivo, fecha, notas, "SCC"])
        documentos_generados += 1

    # --- Hoja Movimientos ---
    ws_mov = wb.create_sheet("Movimientos")
    ws_mov.append(
        [
            "COMPANIA",
            "CENTRO DE OPERACION CO",
            "CONSECUTIVO",
            "NRO MOVIMIENTOS",
            "BODEGA",
            "MOTIVO 01/52",
            "CO DE MOVIMIENTO",
            "CENTRO DE COSTO",
            "UM ORDEN",
            "CANT A SOLICITAR",
            "FECHA AAAAMMDD",
            "CODIGO ITEM",
        ]
    )
    movimientos_generados = 0
    for centro in grupos_orden:
        sub = df[df["C.O. docto."] == centro]
        consecutivo = consecutivo_por_grupo[centro]
        for i, (_, row) in enumerate(sub.iterrows(), start=1):
            ws_mov.append(
                [
                    1,
                    row["C.O. movto."],
                    consecutivo,
                    i,
                    row["Bodega"],
                    row["Motivo"],
                    row["C.O. movto."],
                    row["C.Costo"],
                    row["U.M."],
                    row["Cant. solicitada"],
                    row["Fecha solic."],
                    row["Item"],
                ]
            )
            movimientos_generados += 1

    # --- Hoja Final ---
    ws_final = wb.create_sheet("Final")
    ws_final.append(["COMPANIA"])
    ws_final.append([1])

    # Forzar que los códigos de ítem se traten como texto (para no perder ceros a la izquierda)
    item_col_idx = 12  # CODIGO ITEM es la columna 12 en Movimientos
    for row in ws_mov.iter_rows(min_row=2, min_col=item_col_idx, max_col=item_col_idx):
        for cell in row:
            cell.number_format = "@"

    # Encabezados en negrita
    for ws in (ws_inicial, ws_docs, ws_mov, ws_final):
        for cell in ws[1]:
            cell.font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resumen = {
        "total_filas_entrada": total_filas,
        "descartadas_por_estado": descartadas_estado,
        "documentos_generados": documentos_generados,
        "movimientos_generados": movimientos_generados,
    }
    return buf.read(), resumen
